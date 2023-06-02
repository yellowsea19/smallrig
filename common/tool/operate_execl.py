import openpyxl

def read_excel_data(file_path):
    """
    读取指定路径下的 Excel 表格，将数据转换成字典并放入列表中
    :param file_path: Excel 文件路径
    :return: 转换后的数据列表
    """
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 获取当前活跃的工作表（通常是第一个工作表）
    worksheet = workbook.active

    # 获取表格的行数和列数
    row_count = worksheet.max_row
    column_count = worksheet.max_column

    # 获取表格的表头（即第一行数据），作为字典的 key
    keys = [worksheet.cell(row=1, column=i).value for i in range(1, column_count + 1)]

    # 读取表格的数据，转换成字典，并将字典放入列表中
    data = []
    for i in range(2, row_count + 1):
        row_data = {}
        for j in range(1, column_count + 1):
            row_data[keys[j - 1]] = worksheet.cell(row=i, column=j).value
        data.append(row_data)

    return data

if __name__ == "__main__" :
    data_list = read_excel_data(r"E:\imge\test1.xlsx")
    print(data_list)
